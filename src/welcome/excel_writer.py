"""Append/upsert parsed members into welcome_committee.xlsx."""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
XLSX_PATH = REPO_ROOT / "welcome_committee.xlsx"
TEMPLATE_XLSX_PATH = REPO_ROOT / "welcome_committee_template.xlsx"

HEADERS = ["Date Received", "Family Last Name", "First Name", "Last Name", "Email", "Ministries"]
SHEET_NAME = "Members"
COLUMNS_SHEET_NAME = "Sheet2"
MINISTRY_LEADS_ROW = 2  # User-filled placeholder; preserved across runs.


def _format_date(epoch: int | None) -> str:
    if epoch is None:
        return ""
    return datetime.fromtimestamp(epoch, tz=timezone.utc).astimezone().strftime("%Y-%m-%d")


def _ensure_workbook(path: Path):
    if path.exists():
        wb = load_workbook(path)
        if SHEET_NAME not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_NAME)
            ws.append(HEADERS)
        return wb
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(HEADERS)
    return wb


def upsert_members(
    members_with_ministries: Iterable[tuple[dict, set[str]]],
    path: Path = XLSX_PATH,
) -> tuple[int, int]:
    """Append new rows or update existing ones. Returns (added, updated).

    `members_with_ministries` is an iterable of (member_dict, ministries_set) pairs.
    Members are deduped on lowercased email. On match: the Ministries column is
    replaced with the union of old and new ministries; Date Received keeps whichever
    is earlier.
    """
    wb = _ensure_workbook(path)
    ws = wb[SHEET_NAME]

    email_col = HEADERS.index("Email") + 1
    date_col = HEADERS.index("Date Received") + 1
    ministries_col = HEADERS.index("Ministries") + 1

    existing: dict[str, int] = {}
    for row_idx in range(2, ws.max_row + 1):
        cell_email = ws.cell(row=row_idx, column=email_col).value
        if cell_email:
            existing[str(cell_email).strip().lower()] = row_idx

    added = 0
    updated = 0
    for member, ministries in members_with_ministries:
        email = (member.get("email") or "").strip().lower()
        if not email:
            continue
        new_date = _format_date(member.get("date_received_epoch"))
        new_ministries_set = set(ministries)
        new_ministries_str = ", ".join(sorted(new_ministries_set))

        if email in existing:
            row_idx = existing[email]
            old_ministries_cell = ws.cell(row=row_idx, column=ministries_col).value or ""
            old_set = {m.strip() for m in str(old_ministries_cell).split(",") if m.strip()}
            merged = old_set | new_ministries_set
            ws.cell(row=row_idx, column=ministries_col).value = ", ".join(sorted(merged))

            old_date_cell = ws.cell(row=row_idx, column=date_col).value or ""
            if new_date and (not old_date_cell or new_date < str(old_date_cell)):
                ws.cell(row=row_idx, column=date_col).value = new_date
            updated += 1
        else:
            ws.append(
                [
                    new_date,
                    member.get("family_name", "") or "",
                    member.get("first_name", "") or "",
                    member.get("last_name", "") or "",
                    email,
                    new_ministries_str,
                ]
            )
            existing[email] = ws.max_row
            added += 1

    wb.save(path)
    return added, updated


def write_template_xlsx(
    source: Path = XLSX_PATH,
    dest: Path = TEMPLATE_XLSX_PATH,
) -> None:
    """Build welcome_committee_template.xlsx from welcome_committee.xlsx.

    Two sheets, matching the layout of welcome-committee-example.xlsx:
      * 'Members'  — verbatim copy of the source Members sheet
      * 'Sheet2'   — one column per ministry, ministry name in row 1, ministry-lead
                     email in row 2 (preserved across runs), member emails row 3+.

    Regenerated from scratch on each call. The only data preserved across runs is
    row 2 of Sheet2 (the ministry-lead emails the user types in manually)."""
    # Lazy import to avoid a hard dependency at module-load time on the data_parser
    # constants if the user only runs `welcome bulletin`.
    from .data_parser import MINISTRY_BUCKETS

    if not source.exists():
        raise FileNotFoundError(f"Source workbook not found: {source}")

    src_wb = load_workbook(source)
    src_ws = src_wb[SHEET_NAME]

    preserved_leads: dict[str, str] = {}
    if dest.exists():
        old = load_workbook(dest)
        if COLUMNS_SHEET_NAME in old.sheetnames:
            old_ws = old[COLUMNS_SHEET_NAME]
            for col_idx, ministry in enumerate(MINISTRY_BUCKETS, start=1):
                value = old_ws.cell(row=MINISTRY_LEADS_ROW, column=col_idx).value
                if value:
                    preserved_leads[ministry] = value

    src_headers = [c.value for c in src_ws[1]]
    email_idx = src_headers.index("Email")
    ministries_idx = src_headers.index("Ministries")

    ministry_emails: dict[str, list[str]] = {b: [] for b in MINISTRY_BUCKETS}
    for row in src_ws.iter_rows(min_row=2, values_only=True):
        email = row[email_idx]
        ministries_cell = row[ministries_idx] or ""
        if not email:
            continue
        for m in (s.strip() for s in str(ministries_cell).split(",")):
            if m in ministry_emails:
                ministry_emails[m].append(str(email))

    out = Workbook()
    members_ws = out.active
    members_ws.title = SHEET_NAME
    for row in src_ws.iter_rows(values_only=True):
        members_ws.append(list(row))

    sheet2 = out.create_sheet(COLUMNS_SHEET_NAME)
    sheet2.append(list(MINISTRY_BUCKETS))
    sheet2.append([preserved_leads.get(b, "") for b in MINISTRY_BUCKETS])

    max_rows = max((len(v) for v in ministry_emails.values()), default=0)
    for i in range(max_rows):
        sheet2.append(
            [ministry_emails[b][i] if i < len(ministry_emails[b]) else "" for b in MINISTRY_BUCKETS]
        )

    out.save(dest)
